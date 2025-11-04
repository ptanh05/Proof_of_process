# proof_of_process/ingest/ai_xlsx_ingestor.py
# v3.9 — ONLINE INGEST → STAGING (ổn định, công bằng, “tuần 1..N”)
# - Đọc Google Sheet trực tuyến qua export XLSX (không cần tải về).
# - Chọn sheet thông minh (hoặc ép sheet); phát hiện header bền vững.
# - Map cột VN/EN; suy luận Week → chuẩn 1..N nếu không có tuần hợp lệ.
# - Nhận diện LEAD = người đầu tiên trước dấu + ; , / | & (theo yêu cầu).
# - Khử trùng tên thành viên bằng Jaro–Winkler (canonicalization).
# - Xuất STAGING: external_input.{csv,xlsx}, diagnostics.{json,csv}, preview.html.
# - Schema giữ nguyên để pipeline cũ chạy ổn:
#   Project, Task, Assigned_Member, Task_Leader, Assignees_List, Skills,
#   Effort_Score, Planned_Effort, Actual_Cost, Week, Completion_Status, Depends_On

from __future__ import annotations
import os, io, re, json, argparse, warnings, unicodedata
from typing import List, Tuple, Optional, Any, Dict
import numpy as np
import pandas as pd
import requests

try:
    from openpyxl import load_workbook
except Exception as e:
    raise RuntimeError("Cần cài openpyxl: pip install openpyxl") from e

# ====== HẰNG SỐ & TIỆN ÍCH ======

VI_TRUE = {"x","✓","đã","ok","done","hoàn thành","true","1","yes","completed","xong","finish","100%","hoan thanh"}
TASK_HEADER_HINTS = [
    "task","nhiệm vụ","nhiem vu","mô tả","mo ta","description","công việc","cong viec","title","issue","work item"
]
ASSIGNEE_HINTS = [
    "assigned","assignee","người đảm nhận","nguoi dam nhan","owner","phụ trách","phu trach","nguoi lam","assigned to","người làm"
]
WEEK_HINTS = ["week","tuần","chi tiết tuần","chi tiet tuan","sprint","iteration"]
DONE_HINTS = ["check","done","hoàn thành","hoan thanh","xong","100%"]
STATUS_TEXT_HINTS = ["status","trạng thái","trang thai","giai đoạn","giai doan","progress","tiến độ","tinh trang"]
PLANNED_HINTS = ["planned effort","công sức ước tính","cong suc uoc tinh","estimate","ước tính","effort","sp","story point","complexity"]
COST_HINTS = ["actual cost","chi phí","chi phi","cost","ac","thực chi","budget"]
PROJECT_HINTS = ["project","dự án","du an"]
DEP_HINTS = ["depends","phụ thuộc","phu thuoc","dependency","precedence"]

DAY_PAT = re.compile(r"^(mon|tue|wed|thu|fri|sat|sun|monday|tuesday|wednesday|thursday|friday|saturday|sunday)$", re.I)
NUM_ONLY = re.compile(r"^\d+(\.\d+)?$")

def _clean_text(s) -> str:
    s = str(s or "").strip()
    return re.sub(r"\s+"," ",s)

def _norm(s: Any) -> str:
    s = str(s or "").strip()
    s2 = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"\s+", " ", s2)

def _is_tasky(s: str) -> bool:
    if not s: return False
    if NUM_ONLY.fullmatch(s): return False
    if DAY_PAT.fullmatch(s): return False
    return bool(re.search(r"[A-Za-zÀ-ỹ]", s))

def _safe_float(x, default=np.nan):
    try:
        return float(str(x).replace(",","."))  # thập phân VI
    except:
        return default

def _is_google_sheet(u: str) -> bool:
    return isinstance(u, str) and u.startswith("http") and "docs.google.com/spreadsheets" in u

def _export_xlsx_url(u: str) -> str:
    m = re.search(r"/spreadsheets/d/([^/]+)/", u)
    if not m: return u
    # nếu không có gid → export toàn workbook; có gid → cứ để workbook, ta sẽ pick sheet
    return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"

def _dl_xlsx(u: str) -> bytes:
    r = requests.get(u, timeout=60)
    r.raise_for_status()
    return r.content

# ====== SKILLS (nhẹ, không thay đổi pipeline downstream) ======
# Giữ đúng schema (mảng chuỗi JSON); rule đơn giản để không phụ thuộc sklearn.

SKILL_LEXICON = {
    "Backend":[r"\b(api|rest|graphql|backend|server|database|sql|orm|golang|node|java|python|c#|fastapi|postgres|mysql|mongodb|redis|grpc)\b"],
    "Frontend":[r"\b(ui|ux|react|vue|angular|frontend|html|css|tailwind|next\.js|typescript|vite)\b"],
    "Mobile":[r"\b(android|ios|swift|kotlin|react native|flutter)\b"],
    "DevOps":[r"\b(devops|kubernetes|k8s|docker|ci/?cd|terraform|helm|cloud|aws|gcp|azure)\b"],
    "Security":[r"\b(oauth|jwt|owasp|iam|rbac|waf|encryption|tls|mfa)\b"],
    "Data/AI":[r"\b(nlp|ml|ai|xgboost|pandas|numpy|pytorch|tensorflow|dbt|embedding|retrieval|pipeline|sklearn)\b"],
    "Testing/QA":[r"\b(unit test|integration test|e2e|jest|pytest|cypress|selenium|qa)\b"],
    "Design":[r"\b(figma|wireframe|prototype|mockup|ui kit|branding|layout|typography)\b"],
    "Product/PM":[r"\b(product|roadmap|requirement|spec|prd|user story|sprint|kanban|timeline|okr|kpi|backlog|stakeholder)\b"],
    "General":[r"\b(test|bug|fix|refactor|document|doc|meeting|review|support|deploy|release|ops|maintenance)\b"],
}

def infer_skills(tasks: List[str]) -> List[List[str]]:
    out = []
    for t in tasks:
        t0 = (_clean_text(t)).lower()
        votes = {}
        for sk, pats in SKILL_LEXICON.items():
            v = 0
            for p in pats:
                if re.search(p, t0):
                    v += 1
            if v>0: votes[sk] = v
        ranked = sorted(votes.keys(), key=lambda k: votes[k], reverse=True)
        out.append(ranked[:5] if ranked else ["General"])
    return out

# ====== Jaro–Winkler cho canonicalization tên ======

def jw_similarity(a: str, b: str) -> float:
    a, b = _norm(a), _norm(b)
    if not a or not b:
        return 0.0
    la, lb = len(a), len(b)
    max_dist = max(0, max(la, lb)//2 - 1)
    match = 0
    ha = [False]*la; hb = [False]*lb
    for i in range(la):
        start = max(0, i-max_dist); end = min(i+max_dist+1, lb)
        for j in range(start, end):
            if hb[j]: continue
            if a[i] == b[j]:
                ha[i]=True; hb[j]=True; match+=1; break
    if match == 0: return 0.0
    t = 0; p = 0
    for i in range(la):
        if not ha[i]: continue
        while not hb[p]: p += 1
        if a[i] != b[p]: t += 1
        p += 1
    t//=2
    jaro = (match/la + match/lb + (match - t)/match)/3.0
    prefix = 0
    for i in range(min(4, la, lb)):
        if a[i]==b[i]: prefix+=1
        else: break
    return jaro + 0.1 * prefix * (1 - jaro)

def dedupe_members_probabilistic(names: List[str], threshold: float = 0.92) -> Dict[str, str]:
    uniq=[]
    for n in names:
        n=_clean_text(n)
        if n and n not in uniq: uniq.append(n)
    uniq_sorted = sorted(uniq, key=lambda x: (len(x), x))
    groups=[]; used=set()
    for i,s in enumerate(uniq_sorted):
        if s in used: continue
        group=[s]; used.add(s)
        for t in uniq_sorted[i+1:]:
            if t in used: continue
            if jw_similarity(s,t) >= threshold:
                group.append(t); used.add(t)
        groups.append(group)
    def _canonical(cands: List[str]) -> str:
        def quality(x: str):
            non_alpha = len(re.sub(r"[A-Za-zÀ-ỹ ]", "", x))
            return (abs(len(x)-10), non_alpha, x)
        return sorted(cands, key=quality)[0]
    mapping={}
    for g in groups:
        cano=_canonical(g)
        for m in g: mapping[m]=cano
    return mapping

# ====== Header & Sheet pick ======

def score_ws(ws) -> float:
    score=0.0
    name=(ws.title or "").lower()
    if any(k in name for k in ["timeline","task","nhiệm vụ","project","dự án","du an","work","backlog","week","sprint"]):
        score+=2.0
    max_scan=min(120, ws.max_row)
    for r in range(1, min(8, max_scan)+1):
        vals=[ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        s=" ".join([str(x or "").lower() for x in vals])
        for key in TASK_HEADER_HINTS+ASSIGNEE_HINTS+WEEK_HINTS+STATUS_TEXT_HINTS+PLANNED_HINTS+PROJECT_HINTS+COST_HINTS:
            if key in s: score+=0.5
    return score

def header_candidates(ws, max_scan: int = 200) -> List[Tuple[int, List[str], int]]:
    cands=[]
    for r in range(1, min(ws.max_row, max_scan)+1):
        row=[ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        texts=[str(x or "").strip() for x in row]
        while texts and texts[-1] == "": texts.pop()
        if not any(texts): continue
        s=" ".join([t.lower() for t in texts])
        score=0
        for key in TASK_HEADER_HINTS+ASSIGNEE_HINTS+WEEK_HINTS+STATUS_TEXT_HINTS+PLANNED_HINTS+PROJECT_HINTS+COST_HINTS+DEP_HINTS+DONE_HINTS:
            if key in s: score+=1
        if len(texts)>=3: score+=1
        cands.append((r, texts, score))
    cands.sort(key=lambda x: (x[2], len(x[1])), reverse=True)
    return cands[:10]

def find_header(ws, max_scan=200) -> Tuple[int,List[str]]:
    best=(-1, 1, [])
    for r in range(1, min(ws.max_row, max_scan)+1):
        row=[ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        texts=[str(x or "").strip() for x in row]
        while texts and texts[-1] == "": texts.pop()
        if not any(texts): continue
        s=" ".join([t.lower() for t in texts])
        score=0
        for key in TASK_HEADER_HINTS+ASSIGNEE_HINTS+WEEK_HINTS+STATUS_TEXT_HINTS+PLANNED_HINTS+PROJECT_HINTS+COST_HINTS+DEP_HINTS+DONE_HINTS:
            if key in s: score+=1
        if len(texts)>=3: score+=1
        if score>best[0]:
            best=(score, r, texts)
    return best[1], best[2]

def pick_single_sheet(wb, preferred: Optional[str]=None):
    if preferred and preferred in [ws.title for ws in wb.worksheets]:
        for ws in wb.worksheets:
            if ws.title == preferred: return ws
    scored=sorted([(score_ws(ws), ws) for ws in wb.worksheets], key=lambda x: x[0], reverse=True)
    return scored[0][1]

# ====== Mapping cột ======

def map_cols(cols: List[str]) -> Dict[str, Optional[int]]:
    def f(hints):
        for i,c in enumerate(cols):
            lc=c.lower()
            for k in hints:
                if k in lc: return i
        return None
    return {
        "task": f(TASK_HEADER_HINTS),
        "assignee": f(ASSIGNEE_HINTS),
        "week": f(WEEK_HINTS),
        "done": f(DONE_HINTS),
        "status_text": f(STATUS_TEXT_HINTS),
        "planned": f(PLANNED_HINTS),
        "cost": f(COST_HINTS),
        "project": f(PROJECT_HINTS),
        "depends": f(DEP_HINTS),
    }

# ====== Status & Assignee ======

def detect_done(v) -> bool:
    s=str(v or "").strip().lower()
    if s in VI_TRUE: return True
    if "done" in s or "hoàn thành" in s or "hoan thanh" in s or "completed" in s or "xong" in s: return True
    if "100%" in s: return True
    return False

def normalize_status_value(v: str) -> str:
    s=_clean_text(v).lower()
    if s in ["","0","none","nan","pending","todo","to do","not started","chưa","chua","false","no"]:
        return "Not Started"
    if any(k in s for k in ["doing","in progress","progress","đang làm","dang lam","wip","processing"]):
        return "In Progress"
    if any(k in s for k in ["done","hoàn thành","hoan thanh","completed","complete","finished","xong","100%","ok","check","✓"]):
        return "Completed"
    return "In Progress"

def infer_status_column(df: pd.DataFrame, status_idx: Optional[int]) -> Optional[pd.Series]:
    if status_idx is not None and 0 <= status_idx < len(df.columns):
        col=df.columns[status_idx]
        return df[col].map(normalize_status_value)
    for c in df.columns:
        lc=c.lower()
        if any(k in lc for k in STATUS_TEXT_HINTS):
            return df[c].map(normalize_status_value)
    return None

ASSIG_SPLIT = re.compile(r"[+;,/|&]+")

def assignees_list_from_text(s: str) -> List[str]:
    s=_clean_text(s)
    if not s: return []
    toks=[t.strip() for t in ASSIG_SPLIT.split(s) if t.strip()]
    return toks

def assignees_json(s: str) -> str:
    return json.dumps(assignees_list_from_text(s), ensure_ascii=False)

def choose_lead(assigned_text: str, assignees_json_text: str) -> str:
    # Ưu tiên người đứng trước dấu phân tách (theo yêu cầu)
    lst=assignees_list_from_text(assigned_text)
    if lst: return _clean_text(lst[0])
    try:
        arr=json.loads(assignees_json_text) if assignees_json_text else []
        if isinstance(arr,list) and arr:
            return _clean_text(arr[0])
    except Exception:
        pass
    return ""

# ====== Week ======

def infer_weeks(df: pd.DataFrame, week_col: Optional[str]) -> List[int]:
    # Nếu có cột week (numeric) → dùng
    if week_col and week_col in df.columns:
        w=pd.to_numeric(df[week_col], errors="coerce")
        if w.notna().any():
            w = w.ffill().fillna(1).astype(int).clip(1)

            return w.tolist()
    # Thử từ cột ngày → ISO week
    for c in df.columns:
        if any(k in _norm(c) for k in ["date","ngay","due","deadline","start","end","created","closed","completed","finish"]):
            dt=pd.to_datetime(df[c], errors="coerce", dayfirst=True)
            if dt.notna().sum()>=max(3, int(0.2*len(df))):
                wk=dt.dt.isocalendar().week.astype("Int64")
                wk=wk.fillna(method="ffill").fillna(method="bfill").fillna(1).astype(int)
                # BÓC VỀ 1..N (để biểu đồ không rối Năm/Tháng)
                uniq=[]; remap={}
                for x in wk.tolist():
                    if x not in uniq: uniq.append(int(x))
                for i, val in enumerate(uniq, start=1):
                    remap[val]=i
                return [remap[int(x)] for x in wk.tolist()]
    # Fallback: tuần 1..N theo thứ tự dòng
    return list(range(1, len(df)+1))

def estimate_effort(task: str) -> float:
    t=(task or "").lower()
    base=max(1.0, min(10.0, len(t.split())/6.0))
    bonus=0.0
    for k in ["api","integration","backend","database","complex","research","prototype","design","kubernetes","refactor","migration","security","ml","nlp"]:
        if k in t: bonus+=0.8
    return float(np.clip(base+bonus, 1.0, 10.0))

def choose_task_series(df: pd.DataFrame, task_idx: Optional[int]) -> pd.Series:
    cols=list(df.columns)
    cand=[]
    if task_idx is not None and 0 <= task_idx < len(cols):
        cand.append(cols[task_idx])
    for c in cols:
        lc=c.lower()
        if any(k in lc for k in TASK_HEADER_HINTS):
            if c not in cand: cand.append(c)
    if not cand:
        best=None; best_len=-1.0
        for c in cols:
            try: mlen=df[c].astype(str).str.len().mean()
            except: mlen=-1.0
            if mlen>best_len: best_len=mlen; best=c
        cand=[best] if best else []
    if not cand:
        return pd.Series([""]*len(df))
    s=df[cand[0]].astype(str).copy()
    for p in cand[1:]:
        s=s.where(s.str.strip()!="", df[p].astype(str))
    s=s.map(_clean_text).map(lambda x: x if _is_tasky(x) else "")
    return s

# ====== BUILD FRAME ======
# ====== Skill inference (VN + EN) ======
import re
from collections import defaultdict

# Từ điển từ khóa đa ngôn ngữ (tiếng Việt + tiếng Anh), dùng pattern đơn giản, không phụ thuộc thư viện nặng
SKILL_LEXICON: dict[str, list[str]] = {
    "Frontend": [
        r"\breact\b", r"next\.?js", r"\bvue\b", r"\bangular\b", r"\bsvelte\b",
        r"\b(ui|ux)\b", r"\binterface\b", r"\bcomponent(s)?\b",
        r"\b(css|html|tailwind|responsive|accessibility|aria|storybook)\b",
        r"\bwireframe\b", r"\bprototype\b", r"\bfigma\b", r"\bsketch\b",
        r"giao ?di[eê]n", r"m[aà]n h[iì]nh", r"thi[eế]t k[eế] ui", r"\bfrontend\b",
    ],
    "Backend": [
        r"\b(api|rest|graphql|grpc|websocket|service|microservice)\b",
        r"\b(database|db|postgres|mysql|sqlserver|sqlite|nosql|mongodb|redis|cache|orm)\b",
        r"\b(django|flask|fastapi|spring|laravel|nest\.?js|node\.?js|express)\b",
        r"\b(python|java|golang|go|\.net|c#|rust)\b",
        r"k[eê]t ?n[ôo]i", r"t[íi]ch ?h[ợơ]p", r"x[âa]y d[ựu]ng api",
    ],
    "DevOps": [
        r"\b(deploy|deployment|release|ci/cd|pipeline|jenkins|github actions|gitlab ci)\b",
        r"\b(docker|kubernetes|k8s|helm|terraform|ansible)\b",
        r"\b(cloud|aws|azure|gcp|s3|ec2|eks)\b",
        r"\b(monitor|observability|prometheus|grafana|logging|elastic|elk)\b",
        r"h[ạa] t[ầa]ng", r"tri[eể]n khai", r"v[ậa]n h[àa]nh", r"m[ôo]i tr[uư]ờng",
    ],
    "Security": [
        r"\b(security|auth|oauth|jwt|sso|acl|rbac|iam|tls|cert)\b",
        r"\b(pentest|owasp|xss|csrf|sql injection)\b",
        r"b[aả]o m[ậa]t", r"quy[eề]n truy c[ậa]p", r"x[aá]c th[ựu]c", r"m[ãa] h[óo]a",
    ],
    "Data/AI": [
        r"\b(data|analytics|etl|elt|airflow|dbt|warehouse|lake|bigquery|snowflake|hive|spark|pandas|numpy)\b",
        r"\b(ml|machine learning|deep learning|model|training|inference|sklearn|xgboost|lightgbm|transformer|llm|embedding|prompt|rag)\b",
        r"\b(ab[\s\-]?test|experiment[s]?)\b",
        r"ph[âa]n t[íi]ch d[ữu] li[ệe]u", r"th[íi] ngh[ii]em", r"tr[íi] tu[ệe] nh[âa]n t[ạa]o",
    ],
    "Testing/QA": [
        r"\b(test|unit|integration|e2e|selenium|pytest|jest|cypress|coverage)\b",
        r"\b(benchmark|stress|load|perf(ormance)?)\b",
        r"ki[eể]m th[uử]", r"th[uử] nghi[eệ]m",
    ],
    "Design": [
        r"\b(design|ux|ui|prototype|wireframe|usability|figma|sketch|illustrator|photoshop|branding|styleguide|typography|layout)\b",
        r"thi[eế]t k[eế]", r"tr[ảa]i nghi[eẹ]m ng[ưư]ời d[ùu]ng",
    ],
    "Product/PM": [
        r"\b(product|roadmap|backlog|grooming|sprint|plan(ning)?|estimate|story point|user story|acceptance criteria)\b",
        r"\b(okrs?|kpi|stakeholder|timeline|scope|strategy|business|market|pricing|go(\s|-)to(\s|-)market)\b",
        r"\b(requirement|brd|prd)\b",
        r"qu[ảa]n l[ýy] d[ựu] [áa]n", r"ph[âa]n t[íi]ch th[ịi] tr[ưư]ờng", r"nghi[êe]n c[ứu]", r"k[ếe] ho[ạa]ch",
    ],
    "Mobile": [
        r"\bmobile\b", r"\bandroid\b", r"\bios\b", r"react native", r"\bflutter\b", r"\bkotlin\b", r"\bswift\b",
        r"\bxcode\b", r"\bplay console\b", r"\bapp store\b", r"\bapk\b", r"\bipa\b", r"ứng d[ụu]ng di d[ộo]ng",
    ],
    "Testing/QA": [  # giữ nguyên, đã khai ở trên (đừng xoá)
        r"\b(test|unit|integration|e2e|selenium|pytest|jest|cypress|coverage)\b",
        r"\b(benchmark|stress|load|perf(ormance)?)\b",
        r"ki[eể]m th[uử]", r"th[uử] nghi[eệ]m",
    ],
    "General": [
        r"\b(doc|documentation|meeting|support|bug|fix|refactor|cleanup|research)\b",
        r"h[oọ]p", r"ghi ch[uú]",
    ],
}

def infer_skills(tasks: list[str], topk: int = 4) -> list[list[str]]:
    """
    Suy luận đa nhãn kỹ năng từ mô tả task (VN+EN) theo cơ chế đếm khớp từ khoá
    + vài luật mở rộng đồng xuất hiện. Trả về tối đa topk kỹ năng cho mỗi task.
    """
    out: list[list[str]] = []
    for raw in tasks:
        text = _clean_text(raw).lower()
        if not text:
            out.append(["General"]); continue

        score: dict[str, float] = defaultdict(float)

        # 1) Đếm khớp từ khoá
        for cat, patterns in SKILL_LEXICON.items():
            for pat in patterns:
                try:
                    for _ in re.finditer(pat, text, flags=re.I):
                        score[cat] += 1.0
                except re.error:
                    # pattern lỗi -> bỏ qua
                    continue

        # 2) Luật mở rộng nhẹ theo đồng xuất hiện (giúp không bị “thiếu chiều”)
        if score.get("Frontend", 0) > 0 and any(k in text for k in ["api", "service", "backend"]):
            score["Backend"] += 0.5
        if score.get("Backend", 0) > 0 and any(k in text for k in ["deploy", "docker", "k8s", "triển khai", "trien khai"]):
            score["DevOps"] += 0.5
        if any(k in text for k in ["ab test", "ab-test", "experiment", "thí nghiệm", "thi nghiem"]):
            score["Data/AI"] += 0.5
            score["Product/PM"] += 0.5
        if any(k in text for k in ["bảo mật", "bao mat", "security", "auth", "jwt", "sso"]):
            score["Security"] += 0.75

        # 3) Chọn top-k nhãn có điểm > 0
        ranked = sorted(score.items(), key=lambda kv: kv[1], reverse=True)
        labels = [k for k, v in ranked if v > 0][:topk]

        # 4) Fallback an toàn
        if not labels:
            labels = ["General"]

        out.append(labels)

    return out


def build_frame(sheet_title: str, df: pd.DataFrame, m: dict) -> pd.DataFrame:
    cols=list(df.columns)
    pick=lambda key: (cols[m[key]] if m.get(key) is not None else None)

    col_task=pick("task"); col_ass=pick("assignee"); col_week=pick("week")
    col_done=pick("done"); col_stat=pick("status_text")
    col_plan=pick("planned"); col_cost=pick("cost")
    col_proj=pick("project"); col_dep=pick("depends")

    # Task
    s_task=choose_task_series(df, m.get("task"))
    work=df.copy()
    work["__task__"]=s_task
    work=work[work["__task__"].str.strip()!=""].reset_index(drop=True)
    if work.empty:
        return pd.DataFrame(columns=["Project","Task","Assigned_Member","Task_Leader","Assignees_List","Skills","Effort_Score","Planned_Effort","Actual_Cost","Week","Completion_Status","Depends_On"])

    # Assignee
    if col_ass in work.columns:
        ass_col_series=work[col_ass].astype(str).fillna("")
    else:
        ass_col_series=pd.Series([""]*len(work))
        for c in work.columns:
            if any(k in c.lower() for k in ASSIGNEE_HINTS):
                ass_col_series=work[c].astype(str); break

    ass_json=[assignees_json(a) for a in ass_col_series.tolist()]
    lead_series=pd.Series([choose_lead(a, j) for a, j in zip(ass_col_series.tolist(), ass_json)], index=work.index)

    # Bỏ task không có lead
    keep_mask=lead_series.map(lambda x: bool(x.strip()))
    work=work[keep_mask].copy()
    if work.empty:
        return pd.DataFrame(columns=["Project","Task","Assigned_Member","Task_Leader","Assignees_List","Skills","Effort_Score","Planned_Effort","Actual_Cost","Week","Completion_Status","Depends_On"])
    lead_series=lead_series[keep_mask]
    ass_json=[ass_json[i] for i in keep_mask.index if keep_mask[i]]

    # Status logic
    found_done_source=False
    if col_done in work.columns:
        done_bool=work[col_done].map(detect_done)
        found_done_source=True
    else:
        db=pd.Series([False]*len(work), index=work.index)
        for c in work.columns:
            if any(k in c.lower() for k in DONE_HINTS):
                db=db | work[c].map(detect_done); found_done_source=True
        done_bool=db

    stat_text_series=None
    if not found_done_source:
        if col_stat in work.columns:
            stat_text_series=work[col_stat].map(normalize_status_value)
        else:
            for c in work.columns:
                if any(k in c.lower() for k in STATUS_TEXT_HINTS):
                    stat_text_series=work[c].map(normalize_status_value); break

    if found_done_source:
        completion=["Completed" if b else "Not Started" for b in done_bool.tolist()]
    else:
        if stat_text_series is None:
            completion=["Not Started"]*len(work)
        else:
            uniq=set([x for x in stat_text_series.dropna().unique().tolist() if x])
            if "In Progress" in uniq or len(uniq) >= 3:
                completion=stat_text_series.tolist()
            else:
                completion=["Completed" if s=="Completed" else "Not Started" for s in stat_text_series.tolist()]

    # Week (chuẩn 1..N nếu không suy luận được)
    weeks=infer_weeks(work, col_week)

    # Project
    if col_proj in work.columns:
        projs=work[col_proj].astype(str).map(lambda x: x if x.strip() else sheet_title).tolist()
    else:
        projs=[sheet_title]*len(work)

    # Depends
    deps=work[col_dep].astype(str).fillna("") if col_dep in work.columns else pd.Series([""]*len(work), index=work.index)

    # Effort/Planned/Cost
    eff=work["__task__"].map(estimate_effort)
    if col_plan in work.columns:
        plan=work[col_plan].map(_safe_float)
    else:
        plan=pd.Series([np.nan]*len(work), index=work.index)
    if col_cost in work.columns:
        cost=work[col_cost].map(_safe_float)
    else:
        cost=pd.Series([np.nan]*len(work), index=work.index)

    unit_cost=150.0
    plan=plan.fillna(eff)
    cost=cost.fillna(eff*unit_cost)

    # Skills
    skills=infer_skills(work["__task__"].tolist())
    skills_json=[json.dumps(s, ensure_ascii=False) for s in skills]

    out=pd.DataFrame({
        "Project": projs,
        "Task": work["__task__"].map(_clean_text),
        "Assigned_Member": lead_series.map(_clean_text),   # giữ tương thích pipeline
        "Task_Leader": lead_series.map(_clean_text),
        "Assignees_List": ass_json,
        "Skills": skills_json,
        "Effort_Score": [round(float(x),2) for x in eff.tolist()],
        "Planned_Effort": [round(float(x),2) for x in plan.tolist()],
        "Actual_Cost": [round(float(x),2) for x in cost.tolist()],
        "Week": weeks,
        "Completion_Status": completion,
        "Depends_On": deps.map(_clean_text)
    })

    # Chuẩn hoá & loại trùng (Project, Task, Task_Leader) + canonicalize member
    out=out[out["Task"].str.len()>=2].drop_duplicates(subset=["Project","Task","Task_Leader"]).reset_index(drop=True)
    mapping = dedupe_members_probabilistic(out["Assigned_Member"].astype(str).tolist(), threshold=0.92)
    out["Assigned_Member"] = out["Assigned_Member"].astype(str).map(lambda x: mapping.get(_clean_text(x), _clean_text(x)))
    out["Task_Leader"] = out["Task_Leader"].astype(str).map(lambda x: mapping.get(_clean_text(x), _clean_text(x)))

    # Chuẩn kiểu Week (int >=1) cho biểu đồ tuần 1..N
    out["Week"] = pd.to_numeric(out["Week"], errors="coerce").fillna(1).astype(int).clip(1)

    return out

# ====== INGEST DRIVER (→ thư mục STAGING) ======

def ingest_xlsx(in_path: str, sheet: Optional[str]=None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    if _is_google_sheet(in_path):
        content=_dl_xlsx(_export_xlsx_url(in_path))
        wb=load_workbook(io.BytesIO(content), data_only=True)
    else:
        if not os.path.exists(in_path):
            raise FileNotFoundError(f"Không tìm thấy file: {in_path}")
        wb=load_workbook(in_path, data_only=True)

    ws = pick_single_sheet(wb, preferred=sheet)
    sheet_title = ws.title

    # Header
    hdr_row, hdr_cols = find_header(ws)
    data=[]
    max_cols=len(hdr_cols)
    for r in range(ws.max_row - hdr_row):
        rr=hdr_row + 1 + r
        if rr>ws.max_row: break
        row=[ws.cell(row=rr, column=c).value for c in range(1, max_cols+1)]
        while row and (row[-1] is None or str(row[-1]).strip()==""):
            row.pop()
        if not row: continue
        if len(row) < max_cols: row = row + [""]*(max_cols-len(row))
        data.append(row[:max_cols])

    df=pd.DataFrame(data, columns=[_clean_text(c) for c in hdr_cols]).replace({None:""}).dropna(how="all")
    mapping = map_cols(list(df.columns))
    frame = build_frame(sheet_title, df, mapping)

    diag = {
        "picked_sheet": sheet_title,
        "header_row": int(hdr_row),
        "rows_input": int(len(df)),
        "rows_output": int(len(frame)),
        "note": "Lead = người đầu tiên trước dấu + ; , / | & ; Week = 1..N nếu không có cột tuần hợp lệ."
    }
    return frame, diag

def ingest_xlsx_to_dir(in_path: str, staging_dir: str, sheet: Optional[str]=None, preview_rows: int = 200) -> Dict[str, Any]:
    os.makedirs(staging_dir, exist_ok=True)
    df, diag = ingest_xlsx(in_path, sheet=sheet)

    external_csv = os.path.join(staging_dir, "external_input.csv")
    external_xlsx = os.path.join(staging_dir, "external_input.xlsx")
    preview_html = os.path.join(staging_dir, "preview.html")
    raw_snapshot_csv = os.path.join(staging_dir, "raw_snapshot.csv")  # giả lập snapshot thô = cùng dữ liệu đã làm sạch
    diag_json = os.path.join(staging_dir, "diagnostics.json")
    diag_csv = os.path.join(staging_dir, "diagnostics.csv")

    df.to_csv(external_csv, index=False, encoding="utf-8-sig")
    df.to_excel(external_xlsx, index=False)
    df.to_csv(raw_snapshot_csv, index=False, encoding="utf-8-sig")

    # Preview HTML nhẹ
    try:
        df.head(int(preview_rows)).to_html(preview_html, index=False)
    except Exception:
        pass

    with open(diag_json, "w", encoding="utf-8") as f:
        json.dump(diag, f, ensure_ascii=False, indent=2)
    pd.DataFrame([diag]).to_csv(diag_csv, index=False, encoding="utf-8-sig")

    return {
        "staging_dir": os.path.abspath(staging_dir),
        "external_input_csv": os.path.abspath(external_csv),
        "external_input_xlsx": os.path.abspath(external_xlsx),
        "external_input_parquet": None,
        "preview_html": os.path.abspath(preview_html),
        "raw_snapshot_csv": os.path.abspath(raw_snapshot_csv),
        "diagnostics_json": os.path.abspath(diag_json),
        "diagnostics_csv": os.path.abspath(diag_csv),
        "rows_out": int(len(df)),
        "diagnostics": diag
    }

# ====== CLI ======

if __name__ == "__main__":
    warnings.filterwarnings("ignore", category=FutureWarning)
    ap = argparse.ArgumentParser(description="AI XLSX/GoogleSheet Ingestor → STAGING (v3.9)")
    ap.add_argument("--in", dest="in_path", required=True, help="Đường dẫn .xlsx hoặc URL Google Sheet")
    ap.add_argument("--staging", dest="staging_dir", required=True, help="Thư mục staging để xuất file")
    ap.add_argument("--sheet", dest="sheet", default=None, help="Tên sheet muốn ép dùng (tuỳ chọn)")
    ap.add_argument("--preview-rows", dest="preview_rows", type=int, default=200, help="Số dòng preview HTML")
    args = ap.parse_args()

    meta = ingest_xlsx_to_dir(args.in_path, args.staging_dir, sheet=args.sheet, preview_rows=args.preview_rows)
    print(json.dumps(meta, ensure_ascii=False, indent=2))
